"""Neuromarketing Studio report export utilities.

Exports preserve the distinction between measured inputs, model predictions,
derived visual proxies, and empirical experimental results. Export functions
are deterministic and do not call Gemini or mutate analysis results.
"""

from __future__ import annotations

import csv
import html
import json
from copy import copy
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML


THEME = {
    "primary": "2D2D2D",
    "light": "E5E5E5",
    "accent": "6B46C1",
    "success": "E8F5E9",
    "warning": "FFF3E0",
    "muted": "666666",
}


def _flatten(value: Any, prefix: str = "") -> List[Tuple[str, Any]]:
    if isinstance(value, Mapping):
        flattened: List[Tuple[str, Any]] = []
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            flattened.extend(_flatten(child, child_prefix))
        return flattened
    if isinstance(value, list):
        if not value:
            return [(prefix, "")]
        return [(prefix, json.dumps(value, ensure_ascii=False, default=str))]
    return [(prefix, value)]


def _safe_json(report: Mapping[str, Any]) -> str:
    return json.dumps(report, indent=2, ensure_ascii=False, default=str)


def _evidence_status(report: Mapping[str, Any]) -> str:
    return str(report.get("evidence_status") or report.get("report", {}).get("evidence_status") or "MODEL_PREDICTED")


def write_json_report(report: Mapping[str, Any], output_path: str) -> str:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_safe_json(report) + "\n", encoding="utf-8")
    return str(path)


def write_csv_report(report: Mapping[str, Any], output_path: str) -> str:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _flatten(report)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["metric", "value"])
        for metric, value in rows:
            writer.writerow([metric, value])
    return str(path)


def _metric_rows(report: Mapping[str, Any]) -> List[Tuple[str, Any]]:
    sections = ["mvp_diagnostic", "metrics", "linguistics", "scorecard"]
    rows: List[Tuple[str, Any]] = []
    for section in sections:
        if section in report:
            rows.extend(_flatten(report[section], section))
    return rows


def write_xlsx_report(report: Mapping[str, Any], output_path: str) -> str:
    """Create a client-ready workbook with overview, metrics, variants, and limitations."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    metrics_sheet = workbook.create_sheet("Metrics")
    variants_sheet = workbook.create_sheet("Variants")
    limitations_sheet = workbook.create_sheet("Limitations")

    header_fill = PatternFill("solid", fgColor=THEME["primary"])
    section_fill = PatternFill("solid", fgColor=THEME["light"])
    header_font = Font(name="Georgia", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Georgia", size=18, bold=True, color=THEME["primary"])
    body_font = Font(name="Calibri", size=11, color="000000")
    note_font = Font(name="Calibri", size=10, italic=True, color=THEME["muted"])
    thin = Side(style="thin", color="D1D1D1")

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 3

    overview.merge_cells("B2:F2")
    overview["B2"] = "Neuromarketing Studio — Creative Diagnostics Report"
    overview["B2"].font = title_font
    overview["B2"].alignment = Alignment(vertical="center")
    overview.row_dimensions[2].height = 35
    overview["B4"] = "Evidence status"
    overview["C4"] = _evidence_status(report)
    overview["B5"] = "Generated UTC"
    overview["C5"] = datetime.now(timezone.utc).isoformat()
    overview["B6"] = "Experiment ID"
    overview["C6"] = report.get("experiment_id", "")
    overview["B8"] = "Key metrics"
    overview["B8"].fill = section_fill
    overview.merge_cells("B8:F8")
    overview["B8"].font = Font(name="Georgia", size=13, bold=True, color=THEME["primary"])
    overview["B9"] = "Metric"
    overview["C9"] = "Value"
    for cell in overview[9][1:3]:
        cell.fill = header_fill
        cell.font = header_font
    summary_rows = _metric_rows(report)[:12]
    for row_number, (metric, value) in enumerate(summary_rows, start=10):
        overview.cell(row_number, 2, metric).font = body_font
        overview.cell(row_number, 3, value if isinstance(value, (int, float, str, bool)) else json.dumps(value, default=str)).font = body_font
        overview.cell(row_number, 2).border = Border(bottom=thin)
        overview.cell(row_number, 3).border = Border(bottom=thin)
    overview["B24"] = "Interpretation boundary"
    overview["B24"].fill = section_fill
    overview.merge_cells("B24:F24")
    overview["B24"].font = Font(name="Georgia", size=13, bold=True, color=THEME["primary"])
    overview.merge_cells("B25:F27")
    overview["B25"] = str(report.get("interpretation_boundary") or "Model-derived outputs do not establish participant attention, memory, emotion, neural activity, causal lift, or guaranteed conversion without an appropriate empirical design.")
    overview["B25"].font = note_font
    overview["B25"].alignment = Alignment(wrap_text=True, vertical="top")

    metrics_sheet.append(["Metric", "Value"])
    for cell in metrics_sheet[1][:2]:
        cell.fill = header_fill
        cell.font = header_font
    for metric, value in _metric_rows(report):
        metrics_sheet.append([metric, value if isinstance(value, (int, float, str, bool)) else json.dumps(value, default=str)])
    metrics_sheet.freeze_panes = "B2"
    metrics_sheet.auto_filter.ref = f"B1:C{metrics_sheet.max_row}"
    metrics_sheet.conditional_formatting.add(f"C2:C{metrics_sheet.max_row}", ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color=THEME["accent"]))

    variants_sheet.append(["Variant comparison"])
    variants_sheet["A1"].fill = header_fill
    variants_sheet["A1"].font = header_font
    variants_sheet.append(["Variant comparison is available after a second approved creative is analysed. The MVP does not fabricate a winner or lift estimate."])
    variants_sheet["A2"].alignment = Alignment(wrap_text=True)
    variants_sheet.freeze_panes = "B2"

    limitations = report.get("limitations") or report.get("neuromarketing_indices", {}).get("visual_approach_proxy", {}).get("not_measured", [])
    limitations_sheet.append(["Limitation / boundary"])
    limitations_sheet["A1"].fill = header_fill
    limitations_sheet["A1"].font = header_font
    if isinstance(limitations, list):
        for item in limitations:
            limitations_sheet.append([str(item)])
    else:
        limitations_sheet.append([str(limitations)])
    limitations_sheet.column_dimensions["A"].width = 100
    for row in range(2, limitations_sheet.max_row + 1):
        limitations_sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")

    for sheet in workbook.worksheets:
        for column in range(2, sheet.max_column + 1):
            letter = get_column_letter(column)
            if sheet.column_dimensions[letter].width is None:
                sheet.column_dimensions[letter].width = 22
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and cell.font == Font():
                    cell.font = body_font
                alignment = copy(cell.alignment)
                alignment.vertical = "center"
                cell.alignment = alignment
    workbook.save(path)
    return str(path)


def _html_table(rows: Iterable[Tuple[str, Any]]) -> str:
    body = "".join(
        f"<tr><td>{html.escape(str(metric))}</td><td>{html.escape(str(value))}</td></tr>"
        for metric, value in rows
    )
    return f"<table><thead><tr><th>Metric</th><th>Value</th></tr></thead><tbody>{body}</tbody></table>"


def render_html_report(report: Mapping[str, Any], output_path: str) -> str:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    title = html.escape(str(report.get("experiment_id") or "Creative Diagnostics"))
    evidence = html.escape(_evidence_status(report))
    limitations = report.get("limitations") or report.get("neuromarketing_indices", {}).get("visual_approach_proxy", {}).get("not_measured", [])
    limitations_html = "".join(f"<li>{html.escape(str(item))}</li>" for item in (limitations if isinstance(limitations, list) else [limitations]))
    document = f"""<!doctype html>
<html lang=\"en\"><head><meta charset=\"utf-8\"><title>Neuromarketing Studio — {title}</title>
<style>
body{{font-family:Inter,Arial,sans-serif;color:#202124;margin:40px;line-height:1.45}}
h1{{font-family:Georgia,serif;color:#2D2D2D;border-bottom:4px solid #6B46C1;padding-bottom:10px}}
h2{{font-family:Georgia,serif;color:#2D2D2D;margin-top:28px}}
.badge{{display:inline-block;background:#E5E5E5;padding:6px 12px;border-radius:14px;font-weight:700}}
table{{border-collapse:collapse;width:100%;margin-top:12px}} th{{background:#2D2D2D;color:white;text-align:left}} td,th{{padding:8px;border:1px solid #D1D1D1;vertical-align:top}} tr:nth-child(even){{background:#F7F7F7}}
.note{{background:#FFF3E0;border-left:4px solid #F57C00;padding:12px}}
footer{{color:#666;font-size:0.85em;margin-top:36px}}
</style></head><body>
<h1>Neuromarketing Studio — Creative Diagnostics</h1>
<p><strong>Experiment:</strong> {title} &nbsp; <span class=\"badge\">Evidence: {evidence}</span></p>
<h2>Summary metrics</h2>{_html_table(_metric_rows(report))}
<h2>Interpretation boundary</h2><div class=\"note\">{html.escape(str(report.get("interpretation_boundary") or "These are model-derived visual diagnostics, not observed participant or neural outcomes."))}</div>
<h2>Limitations</h2><ul>{limitations_html}</ul>
<h2>Raw result envelope</h2><pre>{html.escape(_safe_json(report))}</pre>
<footer>Generated by Neuromarketing Studio. Review model assumptions and validate behavioural outcomes with an appropriate empirical study before making causal claims.</footer>
</body></html>"""
    path.write_text(document, encoding="utf-8")
    return str(path)


def write_pdf_report(report: Mapping[str, Any], output_path: str) -> str:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile_html(report) as html_path:
        HTML(filename=html_path).write_pdf(str(path))
    return str(path)


class tempfile_html:
    def __init__(self, report: Mapping[str, Any]):
        import tempfile
        self._temporary = tempfile.NamedTemporaryFile(suffix=".html", delete=False)
        self._temporary.close()
        render_html_report(report, self._temporary.name)

    def __enter__(self) -> str:
        return self._temporary.name

    def __exit__(self, exc_type, exc, tb) -> None:
        try:
            os.unlink(self._temporary.name)
        except FileNotFoundError:
            pass


def export_all(report: Mapping[str, Any], output_dir: str, stem: str = "neuromarketing_report") -> Dict[str, str]:
    directory = Path(output_dir)
    directory.mkdir(parents=True, exist_ok=True)
    return {
        "json": write_json_report(report, str(directory / f"{stem}.json")),
        "csv": write_csv_report(report, str(directory / f"{stem}.csv")),
        "xlsx": write_xlsx_report(report, str(directory / f"{stem}.xlsx")),
        "html": render_html_report(report, str(directory / f"{stem}.html")),
        "pdf": write_pdf_report(report, str(directory / f"{stem}.pdf")),
    }


def validate_exports(paths: Mapping[str, str]) -> Dict[str, Any]:
    """Perform lightweight format validation without interpreting report content."""
    results: Dict[str, Any] = {}
    for kind, filename in paths.items():
        path = Path(filename)
        valid = path.exists() and path.stat().st_size > 0
        if valid and kind == "json":
            json.loads(path.read_text(encoding="utf-8"))
        elif valid and kind == "csv":
            with path.open(newline="", encoding="utf-8") as handle:
                valid = len(list(csv.reader(handle))) >= 2
        elif valid and kind == "xlsx":
            workbook = load_workbook(path, read_only=True)
            valid = "Overview" in workbook.sheetnames and "Metrics" in workbook.sheetnames
            workbook.close()
        elif valid and kind == "pdf":
            valid = path.read_bytes()[:5] == b"%PDF-"
        elif valid and kind == "html":
            valid = "Neuromarketing Studio" in path.read_text(encoding="utf-8")
        results[kind] = {"path": str(path), "valid": bool(valid), "bytes": path.stat().st_size if path.exists() else 0}
    return results
